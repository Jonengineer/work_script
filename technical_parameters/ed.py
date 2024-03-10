import openpyxl

# Загрузка рабочей книги Excel
workbook = openpyxl.load_workbook('ed.xlsx')
sheet = workbook.active  # Предполагаем, что работаем с активной вкладкой

# Перебор всех строк
for row in sheet.iter_rows(min_row=2):  # Пропускаем заголовок
    code = row[2].value  # Столбец 'code'
    if code and len(code) >= 2:
        short_code = code[:2]  # Берем первые два символа из 'code'
        for search_row in sheet.iter_rows(min_row=2):
            uncCode = search_row[0].value  # Столбец 'uncCode'
            if uncCode and uncCode.startswith(short_code):
                row[3].value = search_row[1].value # Присвоение значения из 'unit' в 'unit2'
                break # Прекращаем поиск после первого совпадения

workbook.save('ed.xlsx')