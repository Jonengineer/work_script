import openpyxl
import json

# Открываем файл Excel
file_path = 'test.xlsx'
workbook = openpyxl.load_workbook(file_path)

# Выбираем первую вкладку
target_sheet = workbook['1']  # Замените на имя первой вкладки

# Выбираем исходную вторую вкладку
source_sheet = workbook['2']  # Замените на имя второй вкладки

# Создаем словарь для хранения данных из столбца J второй вкладки
source_data = {}
source_data_2 = {}
for row in source_sheet.iter_rows(min_row=2):
    identifier = row[4].value  # E столбец - 5-й столбец
    j_cell = row[9].value  # J столбец - 10-й столбец
    techParameters = row[6].value 
    techName = row[7].value 
    techDescription = row[8].value 

    if identifier and j_cell:
        try:
            data_dict = json.loads(j_cell)
            source_data[identifier] = data_dict            
        except json.JSONDecodeError:
            print(f"Ошибка в строке {row[0].row}: неверный формат JSON")

    if identifier and (techParameters or techName or techDescription):
        source_data_2[identifier] = {
            'techParameters': techParameters,
            'techName': techName,
            'techDescription': techDescription
        }


# Обновляем столбцы O, P и BD первой вкладки на основе данных из второй вкладки
for row in target_sheet.iter_rows(min_row=2):
    identifier = row[0].value  # A столбец - 1-й столбец

    if identifier in source_data_2:
        data = source_data_2[identifier]
        techName = data.get('techName', '')  # Получение techName, если оно существует, иначе пустая строка
        techParameters = data.get('techParameters', '')  # Аналогично для techParameters
        techDescription = data.get('techDescription', '')  # Аналогично для techDescription

        row[11].value = techName  # O столбец - 15-й столбец
        row[109].value = techParameters  # P столбец - 16-й столбец
        row[110].value = techDescription  # BD столбец - 56-й столбец



    if identifier in source_data:
        json_data = source_data[identifier]

        # Обрабатываем ключ "1 напряжение"
        if "1" in json_data:            
            values = json_data["1"]
            Umin_value = values[0] if len(values) > 0 else None
            Umax_value = values[1] if len(values) > 1 else None
            row[14].value = Umin_value  # O столбец - 15-й столбец
            row[15].value = Umax_value  # P столбец - 16-й столбец

        # Обрабатываем ключ "2 номинальный ток выключателя"
        if "2" in json_data:
            Inom_min = json_data["2"][0] if len(json_data["2"]) > 0 else None
            Inom_max = json_data["2"][1] if len(json_data["2"]) > 1 else None
            
            if Inom_max is None:
                row[17].value = Inom_min
            else:
                row[17].value = Inom_max
  
            row[16].value = Inom_min  

        # Обрабатываем ключ "3 ток отключения выключателя"
        if "3" in json_data:
            Iotkl_min = json_data["3"][0] if len(json_data["3"]) > 0 else None
            Iotkl_max = json_data["3"][1] if len(json_data["3"]) > 1 else None
            
            if Inom_max is None:
                row[19].value = Iotkl_min
            else:
                row[19].value = Iotkl_max
  
            row[18].value = Iotkl_min  

        # Обрабатываем ключ "4 номинальная мощность трансформатора"
        if "4" in json_data:
            Strans_min = json_data["4"][0] if len(json_data["4"]) > 0 else None
            Strans_max = json_data["4"][1] if len(json_data["4"]) > 1 else None
            
            if Strans_max is None:
                row[21].value = Strans_min
            else:
                row[21].value = Strans_max
  
            row[20].value = Strans_min

        # Обрабатываем ключ "5 номинальная тип трансформатора"
        if "5" in json_data:
            Transformer = json_data["5"][0] if len(json_data["5"]) > 0 else None 
            row[104].value = Transformer

        # Обрабатываем ключ "5 номинальная тип трансформатора"
        if "6" in json_data:
            Transformer = json_data["6"][0] if len(json_data["6"]) > 0 else None 
            row[104].value = Transformer
        
        # Обрабатываем ключ "5 тип трансформатора"
        if "7" in json_data:
            Transformer = json_data["7"][0] if len(json_data["7"]) > 0 else None 
            row[104].value = Transformer

        # Обрабатываем ключ "8 сопротивление реактора"
        if "8" in json_data:
            R_min = json_data["8"][0] if len(json_data["8"]) > 0 else None
            R_max = json_data["8"][1] if len(json_data["8"]) > 1 else None
            
            if R_max is None:
                row[29].value = R_min
            else:
                row[29].value = R_max
  
            row[28].value = R_min

        # Обрабатываем ключ "9 мощность реактора"
        if "9" in json_data:
            Sreact_min = json_data["9"][0] if len(json_data["9"]) > 0 else None
            Sreact_max = json_data["9"][1] if len(json_data["9"]) > 1 else None
            
            if Sreact_max is None:
                row[25].value = Sreact_min
            else:
                row[25].value = Sreact_max
  
            row[24].value = Sreact_min

        # Обрабатываем ключ "10 мощность ДРГ"
        if "10" in json_data:
            Sdgr_min = json_data["10"][0] if len(json_data["10"]) > 0 else None
            Sdgr_max = json_data["10"][1] if len(json_data["10"]) > 1 else None
            
            if Sdgr_max is None:
                row[23].value = Sdgr_min
            else:
                row[23].value = Sdgr_max
  
            row[22].value = Sdgr_min

        # Обрабатываем ключ "12 сечение кабеля"
        if "12" in json_data:
            veinsize_min = json_data["12"][0] if len(json_data["12"]) > 0 else None 
            veinsize_max = json_data["12"][1] if len(json_data["12"]) > 1 else None
            
            if veinsize_max is None:
                row[88].value = veinsize_min
            else:
                row[88].value = veinsize_max

            row[87].value = veinsize_min
        
        # Обрабатываем ключ "13 Количество волокон"
        if "13" in json_data:
            qfiber_min = json_data["13"][0] if len(json_data["13"]) > 0 else None 
            qfiber_max = json_data["13"][1] if len(json_data["13"]) > 1 else None
            
            if qfiber_max is None:
                row[67].value = qfiber_min
            else:
                row[67].value = qfiber_max
  
            row[66].value = qfiber_min    

        # Обрабатываем ключ "14 шкафы оборудования"
        if "14" in json_data:
            boxkit = json_data["14"][0] if len(json_data["14"]) > 0 else None 
            row[40].value = boxkit

        # Обрабатываем ключ "15 Количество трансорматоров в ктп"
        if "15" in json_data:
            qtrans_min = json_data["15"][0] if len(json_data["15"]) > 0 else None 
            qtrans_max = json_data["15"][1] if len(json_data["15"]) > 1 else None
            
            if qtrans_max is None:
                row[31].value = qtrans_min
            else:
                row[31].value = qtrans_max
  
            row[30].value = qtrans_min  

        # Обрабатываем ключ "17 шкафы оборудования"
        if "17" in json_data:
            qconnect_min = json_data["17"][0] if len(json_data["17"]) > 0 else None 
            qconnect_max = json_data["17"][1] if len(json_data["17"]) > 1 else None
            
            if qconnect_max is None:
                row[106].value = qconnect_min
            else:
                row[106].value = qconnect_max
  
            row[105].value = qconnect_min

        # Обрабатываем ключ "18 устройство дорожного покрытия"
        if "18" in json_data:
            roadfix = json_data["18"][0] if len(json_data["18"]) > 0 else None 
            row[47].value = roadfix

        # Обрабатываем ключ "19 количество жил"
        if "19" in json_data:
            qvein_min = json_data["19"][0] if len(json_data["19"]) > 0 else None 
            qvein_max = json_data["19"][1] if len(json_data["19"]) > 1 else None
            
            if qvein_max is None:
                row[92].value = qvein_min
            else:
                row[92].value = qvein_max
  
            row[91].value = qvein_min

        # Обрабатываем ключ "20 еденица измерения"
        if "20" in json_data:
            unit_of_measure = json_data["20"][0] if len(json_data["20"]) > 0 else None
            row[10].value = unit_of_measure  # BD столбец - 56-й столбец
        
        # Обрабатываем ключ "21 тип жилы"
        if "21" in json_data:
            veintype = json_data["21"][0] if len(json_data["21"]) > 0 else None
            row[86].value = veintype  # BD столбец - 56-й столбец

        # Обрабатываем ключ "22 шкафы оборудования"
        if "22" in json_data:
            shieldsize_min = json_data["22"][0] if len(json_data["22"]) > 0 else None 
            shieldsize_max = json_data["22"][1] if len(json_data["22"]) > 1 else None
            
            if shieldsize_max is None:
                row[90].value = shieldsize_min
            else:
                row[90].value = shieldsize_max
  
            row[89].value = shieldsize_min

        # Обрабатываем ключ "24 диаметр труб"
        if "24" in json_data:
            dpipe_min = json_data["24"][0] if len(json_data["24"]) > 0 else None 
            dpipe_max = json_data["24"][1] if len(json_data["24"]) > 1 else None
            
            if dpipe_max is None:
                row[102].value = dpipe_min
            else:
                row[102].value = dpipe_max
  
            row[101].value = dpipe_min

        # Обрабатываем ключ "25 тип опор"
        if "25" in json_data:
            opora_type = json_data["25"][0] if len(json_data["25"]) > 0 else None
            row[52].value = opora_type  # BD столбец - 56-й столбец

        # Обрабатываем ключ "26 сечение провода"
        if "26" in json_data:
            Seh_provod_min = json_data["26"][0] if len(json_data["26"]) > 0 else None
            Seh_provod_max = json_data["26"][1] if len(json_data["26"]) > 1 else None
            # Если Seh_0_provod_max равно 0, записываем Seh_0_provod_min в 62-й столбец
            if Seh_provod_max is None:
                row[59].value = Seh_provod_min
            else:
                row[59].value = Seh_provod_max
  
            row[58].value = Seh_provod_min  # Столбец BI - 61-й столбец    

        # Обрабатываем ключ "28 количество проводов в фазе провода"
        if "28" in json_data:
            Count_provod_min = json_data["28"][0] if len(json_data["28"]) > 0 else None
            Count_provod_max = json_data["28"][1] if len(json_data["28"]) > 1 else None
            # Если Seh_0_provod_max равно 0, записываем Seh_0_provod_min в 62-й столбец
            if Count_provod_max is None:
                row[57].value = Count_provod_min
            else:
                row[57].value = Count_provod_max
  
            row[56].value = Count_provod_min  # Столбец BI - 61-й столбец 
        
        # Обрабатываем ключ "29 диаметр грозотроса"
        if "29" in json_data:
            d_rope_min = json_data["29"][0] if len(json_data["29"]) > 0 else None
            d_rope_max = json_data["29"][1] if len(json_data["29"]) > 1 else None
            # Если Seh_0_provod_max равно 0, записываем Seh_0_provod_min в 62-й столбец
            if d_rope_max is None:
                row[65].value = d_rope_min
            else:
                row[65].value = d_rope_max
  
            row[64].value = d_rope_min  # Столбец BI - 61-й столбец 

        # Обрабатываем ключ "30 сечение нулевого провода"
        if "30" in json_data:
            Seh_0_provod_min = json_data["30"][0] if len(json_data["30"]) > 0 else None
            Seh_0_provod_max = json_data["30"][1] if len(json_data["30"]) > 1 else None
            # Если Seh_0_provod_max равно 0, записываем Seh_0_provod_min в 62-й столбец
            if Seh_0_provod_max is None:
                row[61].value = Seh_0_provod_min
            else:
                row[61].value = Seh_0_provod_max
  
            row[60].value = Seh_0_provod_min  # Столбец BI - 61-й столбец
        
        # Обрабатываем ключ "31 тип провода"
        if "31" in json_data:
            Type_provod_value = json_data["31"][0] if len(json_data["31"]) > 0 else None
            row[55].value = Type_provod_value  # BD столбец - 56-й столбец
            print(Type_provod_value)

        # Обрабатываем ключ "32 допустимый длительный ток"
        if "32" in json_data:
            ilong_min = json_data["32"][0] if len(json_data["32"]) > 0 else None
            ilong_max = json_data["32"][1] if len(json_data["32"]) > 1 else None
            # Если Seh_0_provod_max равно 0, записываем Seh_0_provod_min в 62-й столбец
            if ilong_max is None:
                row[63].value = ilong_min
            else:
                row[63].value = ilong_max
  
            row[62].value = ilong_min  # Столбец BI - 61-й столбец
   
        # Обрабатываем ключ "33 количество цепей"
        if "33" in json_data:
            qcircuit_min = json_data["33"][0] if len(json_data["33"]) > 0 else None
            qcircuit_max = json_data["33"][1] if len(json_data["33"]) > 1 else None
            # Если Seh_0_provod_max равно 0, записываем Seh_0_provod_min в 62-й столбец
            if qcircuit_max is None:
                row[54].value = qcircuit_min
            else:
                row[54].value = qcircuit_max
  
            row[53].value = qcircuit_min  # Столбец BI - 61-й столбец

        # Обрабатываем ключ "35 диаметр трубопровода"
        if "35" in json_data:
            dpipeline_min = json_data["35"][0] if len(json_data["35"]) > 0 else None
            dpipeline_max = json_data["35"][1] if len(json_data["35"]) > 1 else None
            # Если Seh_0_provod_max равно 0, записываем Seh_0_provod_min в 62-й столбец
            if dpipeline_max is None:
                row[75].value = dpipeline_min
            else:
                row[75].value = dpipeline_max
  
            row[74].value = dpipeline_min  # Столбец BI - 61-й столбец

        # Обрабатываем ключ "36 механическая прочность"
        if "36" in json_data:
            prochnost_min = json_data["36"][0] if len(json_data["36"]) > 0 else None
            prochnost_max = json_data["36"][1] if len(json_data["36"]) > 1 else None
            # Если Seh_0_provod_max равно 0, записываем Seh_0_provod_min в 62-й столбец
            if prochnost_max is None:
                row[69].value = prochnost_min
            else:
                row[69].value = prochnost_max
  
            row[68].value = prochnost_min  # Столбец BI - 61-й столбец    

        # Обрабатываем ключ "37 максимально-допустимая растягивающая нагрузка"
        if "37" in json_data:
            stretch_min = json_data["37"][0] if len(json_data["37"]) > 0 else None
            stretch_max = json_data["37"][1] if len(json_data["37"]) > 1 else None
            # Если Seh_0_provod_max равно 0, записываем Seh_0_provod_min в 62-й столбец
            if stretch_max is None:
                row[71].value = stretch_min
            else:
                row[71].value = stretch_max
  
            row[70].value = stretch_min  # Столбец BI - 61-й столбец 

        # Обрабатываем ключ "39 протяженность лэеп П8"
        if "39" in json_data:
            lepdist_min = json_data["39"][0] if len(json_data["39"]) > 0 else None
            lepdist_max = json_data["39"][1] if len(json_data["39"]) > 1 else None
            # Если Seh_0_provod_max равно 0, записываем Seh_0_provod_min в 62-й столбец
            if lepdist_max is None:
                row[82].value = lepdist_min
            else:
                row[82].value = lepdist_max
  
            row[81].value = lepdist_min  # Столбец BI - 61-й столбец 
    
        # Обрабатываем ключ "вид объекта"
        if "41" in json_data:
            Type_provod_value = json_data["41"][0] if len(json_data["41"]) > 0 else None
            row[103].value = Type_provod_value  # BD столбец - 56-й столбец

        # Обрабатываем ключ "42 количество землепользователей"
        if "42" in json_data:
            qholders_min = json_data["42"][0] if len(json_data["42"]) > 0 else None
            qholders_max = json_data["42"][1] if len(json_data["42"]) > 1 else None
            # Если Seh_0_provod_max равно 0, записываем Seh_0_provod_min в 62-й столбец
            if qholders_max is None:
                row[80].value = qholders_min
            else:
                row[80].value = qholders_max
  
            row[79].value = qholders_min  # Столбец BI - 61-й столбец 

        # Обрабатываем ключ "43 тип Р4"
        if "43" in json_data:
            type_P4 = json_data["43"][0] if len(json_data["43"]) > 0 else None
            row[107].value = type_P4  # BD столбец - 56-й столбец

        # Обрабатываем ключ "44 регион П9"
        if "44" in json_data:
            region_P9 = json_data["44"][0] if len(json_data["44"]) > 0 else None
            row[108].value = region_P9  # BD столбец - 56-й столбец

# Сохраняем изменения
workbook.save(file_path)